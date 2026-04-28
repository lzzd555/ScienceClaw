from datetime import datetime
from io import BytesIO
from typing import Iterable

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from auth import get_current_user
from database import get_db
from fixtures import write_placeholder_download
from models import Contract, DownloadEvent, PurchaseOrder, ReportJob, User
from schemas import DownloadEventOut, ReportJobOut


REPORT_JOB_ID = "RPT-2026-RPA-001"
REPORT_FILENAME = "supplier_purchase_summary_2026.xlsx"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

router = APIRouter(dependencies=[Depends(get_current_user)])


def workbook_response(workbook: Workbook, filename: str) -> StreamingResponse:
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(stream, media_type=XLSX_MEDIA_TYPE, headers=headers)


def record_download(db: Session, *, filename: str, source: str) -> None:
    db.add(DownloadEvent(filename=filename, source=source))
    db.commit()


def build_contracts_workbook(contracts: Iterable[Contract]) -> Workbook:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("合同台账")
    sheet.append(["合同编号", "合同名称", "供应商编号", "供应商名称", "类型", "状态", "金额", "归口部门"])
    for contract in contracts:
        sheet.append(
            [
                contract.number,
                contract.title,
                contract.supplier.number,
                contract.supplier.name,
                contract.contract_type,
                contract.status,
                contract.amount,
                contract.owner_department,
            ]
        )
    return workbook


def build_purchase_orders_workbook(orders: Iterable[PurchaseOrder]) -> Workbook:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("采购订单")
    sheet.append(["采购订单号", "采购申请号", "供应商编号", "供应商名称", "状态", "优先级", "金额"])
    for order in orders:
        sheet.append(
            [
                order.number,
                order.purchase_request.number,
                order.supplier.number,
                order.supplier.name,
                order.status,
                order.priority,
                order.total_amount,
            ]
        )
    return workbook


def build_merged_workbook(contracts: Iterable[Contract], orders: Iterable[PurchaseOrder]) -> Workbook:
    workbook = Workbook(write_only=True)
    contract_sheet = workbook.create_sheet("合同台账")
    contract_sheet.append(["合同编号", "合同名称", "供应商名称", "类型", "状态", "金额"])
    for contract in contracts:
        contract_sheet.append(
            [
                contract.number,
                contract.title,
                contract.supplier.name,
                contract.contract_type,
                contract.status,
                contract.amount,
            ]
        )

    order_sheet = workbook.create_sheet("采购订单")
    order_sheet.append(["采购订单号", "采购申请号", "供应商名称", "状态", "优先级", "金额"])
    for order in orders:
        order_sheet.append(
            [
                order.number,
                order.purchase_request.number,
                order.supplier.name,
                order.status,
                order.priority,
                order.total_amount,
            ]
        )
    return workbook


def get_or_create_report_job(db: Session) -> ReportJob:
    job = db.query(ReportJob).filter(ReportJob.job_id == REPORT_JOB_ID).one_or_none()
    if job is None:
        job = ReportJob(
            job_id=REPORT_JOB_ID,
            report_type="supplier_purchase_summary",
            status="not_started",
            filename=REPORT_FILENAME,
        )
        db.add(job)
        db.flush()
    return job


def advance_report_job(job: ReportJob) -> None:
    if job.status != "processing":
        return
    job.poll_count += 1
    if job.poll_count >= 2:
        job.status = "completed"
        job.completed_at = datetime.utcnow()
        write_placeholder_download(job.filename)


@router.get("", response_model=list[ReportJobOut])
def list_reports(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> list[ReportJob]:
    return db.query(ReportJob).order_by(ReportJob.job_id).all()


@router.get("/contracts/export")
def export_contracts(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StreamingResponse:
    contracts = db.query(Contract).order_by(Contract.number).all()
    workbook = build_contracts_workbook(contracts)
    record_download(db, filename="contracts_2026.xlsx", source="contracts_export")
    return workbook_response(workbook, "contracts_2026.xlsx")


@router.get("/purchase-orders/export")
def export_purchase_orders(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StreamingResponse:
    orders = db.query(PurchaseOrder).order_by(PurchaseOrder.number).all()
    workbook = build_purchase_orders_workbook(orders)
    record_download(db, filename="purchase_orders_2026.xlsx", source="purchase_orders_export")
    return workbook_response(workbook, "purchase_orders_2026.xlsx")


@router.post("/merge/export")
def export_merged_report(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StreamingResponse:
    contracts = db.query(Contract).order_by(Contract.number).all()
    orders = db.query(PurchaseOrder).order_by(PurchaseOrder.number).all()
    workbook = build_merged_workbook(contracts, orders)
    record_download(db, filename="rpa_eval_merged_export_2026.xlsx", source="merged_export")
    return workbook_response(workbook, "rpa_eval_merged_export_2026.xlsx")


@router.post("/async/generate", response_model=ReportJobOut)
def generate_async_report(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> ReportJob:
    job = get_or_create_report_job(db)
    job.status = "processing"
    job.poll_count = 0
    job.completed_at = None
    job.filename = REPORT_FILENAME
    db.commit()
    db.refresh(job)
    return job


@router.get("/async/{report_code}", response_model=ReportJobOut)
def get_async_report(
    report_code: str,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> ReportJob:
    if report_code != REPORT_JOB_ID:
        raise HTTPException(status_code=404, detail="Report job not found")
    job = get_or_create_report_job(db)
    advance_report_job(job)
    db.commit()
    db.refresh(job)
    return job


@router.get("/async/{report_code}/download")
def download_async_report(
    report_code: str,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StreamingResponse:
    if report_code != REPORT_JOB_ID:
        raise HTTPException(status_code=404, detail="Report job not found")
    job = get_or_create_report_job(db)
    if job.status != "completed":
        raise HTTPException(status_code=409, detail="Report is not completed")

    orders = db.query(PurchaseOrder).order_by(PurchaseOrder.number).all()
    workbook = build_purchase_orders_workbook(orders)
    record_download(db, filename=REPORT_FILENAME, source="async_supplier_purchase_summary")
    return workbook_response(workbook, REPORT_FILENAME)


@router.post("/supplier-purchase-summary/generate", response_model=ReportJobOut)
def generate_supplier_purchase_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ReportJob:
    return generate_async_report(db=db, _=current_user)


@router.get("/download-events", response_model=list[DownloadEventOut])
def list_download_events(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> list[DownloadEvent]:
    return db.query(DownloadEvent).order_by(DownloadEvent.created_at).all()
