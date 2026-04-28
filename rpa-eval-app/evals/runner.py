from __future__ import annotations

import argparse
import json
import re
import sys
import time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlparse

import yaml

if __package__ in (None, ""):
    sys.path.append(str(Path(__file__).resolve().parent))
    from eval_app_client import EvalAppClient, EvalAppError
    from report import summarize_cases, utc_run_id, write_reports
    from rpa_client import RpaClawClient, RpaClawError, RpaClawTimeoutError, RpaRunResult
else:
    from .eval_app_client import EvalAppClient, EvalAppError
    from .report import summarize_cases, utc_run_id, write_reports
    from .rpa_client import RpaClawClient, RpaClawError, RpaClawTimeoutError, RpaRunResult


DEFAULT_REPORT_DIR = Path(__file__).resolve().parent / "reports"
CASES_DIR = Path(__file__).resolve().parent / "cases"
USER_PASSWORDS = {
    "admin": "admin123",
    "buyer": "buyer123",
    "approver": "approver123",
}


class CaseAssertionError(AssertionError):
    def __init__(self, stage: str, message: str) -> None:
        super().__init__(message)
        self.stage = stage


def main() -> int:
    args = parse_args()
    report_dir = args.report_dir if args.report_dir is not None else DEFAULT_REPORT_DIR
    cases = select_cases(load_cases(CASES_DIR), args)
    if not cases:
        print("No eval cases matched the requested selector.", file=sys.stderr)
        return 2

    eval_client = EvalAppClient(args.eval_backend_url)
    rpa_client = RpaClawClient(
        args.rpaclaw_url,
        token=args.rpaclaw_token,
        model_name=args.model,
        model_config_id=args.model_config_id,
    )
    run_id = utc_run_id()
    print(f"Running {len(cases)} RPA eval case(s)")
    if args.model:
        print(f"Model: {args.model} -> {rpa_client.model_config_id}")
    elif args.model_config_id:
        print(f"Model config id: {args.model_config_id}")
    case_results = []
    for index, case in enumerate(cases, start=1):
        timeout_s = resolve_case_timeout_s(case, args)
        print(
            f"[{index}/{len(cases)}] START {case['id']} - {case.get('name', case['id'])} "
            f"(timeout {timeout_s:g}s)",
            flush=True,
        )
        result = run_case(case, args, eval_client, rpa_client)
        case_results.append(result)
        status = "PASS" if result["passed"] else "FAIL"
        detail = "" if result["passed"] else f" {result.get('failure_stage')}: {result.get('failure_message')}"
        print(f"[{index}/{len(cases)}] {status} {case['id']} ({result.get('latency_ms', 0)} ms){detail}", flush=True)
    report = {
        "run_id": run_id,
        "created_at": run_id,
        "config": {
            "eval_backend_url": args.eval_backend_url,
            "eval_frontend_url": args.eval_frontend_url,
            "rpaclaw_url": args.rpaclaw_url,
            "model": args.model,
            "model_config_id": rpa_client.model_config_id,
            "default_case_timeout_s": args.case_timeout_s,
            "selectors": {"case": args.case, "tag": args.tag, "all": args.all},
            "report_dir": str(report_dir),
        },
        "summary": summarize_cases(case_results),
        "cases": case_results,
    }
    paths = write_reports(report, report_dir)
    print()
    print(render_console_summary(case_results))
    print()
    print(f"Wrote JSON report: {paths['latest_json']}")
    print(f"Wrote Markdown report: {paths['latest_md']}")
    return 0 if report["summary"]["failed"] == 0 else 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run RPA golden evaluation cases.")
    parser.add_argument("--case", action="append", default=[], help="Case id to run. Can be repeated.")
    parser.add_argument("--tag", action="append", default=[], help="Tag to run. Can be repeated.")
    parser.add_argument("--all", action="store_true", help="Run all cases.")
    parser.add_argument("--eval-backend-url", default="http://localhost:8085")
    parser.add_argument("--eval-frontend-url", default="http://localhost:5175")
    parser.add_argument("--rpaclaw-url", default="http://localhost:12001")
    parser.add_argument("--rpaclaw-token", default="")
    parser.add_argument("--model", default="", help="RpaClaw model name to use, matched against /api/v1/models.")
    parser.add_argument("--model-config-id", default="", help="RpaClaw model config id to use directly.")
    parser.add_argument("--report-dir", default=None)
    parser.add_argument("--reset-token", default="rpa-eval-reset")
    parser.add_argument(
        "--case-timeout-s",
        type=float,
        default=180.0,
        help="Default wall-clock timeout for one eval case. A case can override it with timeout_s in YAML.",
    )
    return parser.parse_args()


def load_cases(cases_dir: Path) -> list[dict[str, Any]]:
    cases = []
    for path in sorted(cases_dir.glob("*.yaml")):
        with path.open("r", encoding="utf-8") as handle:
            case = yaml.safe_load(handle)
        if not isinstance(case, dict) or not case.get("id"):
            raise ValueError(f"Invalid eval case file: {path}")
        case["_path"] = str(path)
        cases.append(case)
    return cases


def select_cases(cases: list[dict[str, Any]], args: argparse.Namespace) -> list[dict[str, Any]]:
    if args.all:
        return cases
    selected_ids = set(args.case)
    selected_tags = set(args.tag)
    if selected_ids or selected_tags:
        return [
            case
            for case in cases
            if case["id"] in selected_ids or selected_tags.intersection(set(case.get("tags", [])))
        ]
    return [case for case in cases if "smoke" in case.get("tags", [])]


def run_case(
    case: dict[str, Any],
    args: argparse.Namespace,
    eval_client: EvalAppClient,
    rpa_client: RpaClawClient,
) -> dict[str, Any]:
    started = time.perf_counter()
    result: dict[str, Any] = {
        "id": case["id"],
        "name": case.get("name", case["id"]),
        "tags": case.get("tags", []),
        "passed": False,
        "attempts": 1,
        "failure_stage": None,
        "failure_message": None,
        "expected": case.get("expected", {}),
        "metrics": {},
        "raw_events": [],
        "session_id": None,
        "timeout_s": resolve_case_timeout_s(case, args),
    }

    try:
        eval_client.reset(args.reset_token)
        user = case.get("user") or {}
        username = user["username"]
        password = USER_PASSWORDS[username]
        eval_session = eval_client.login(username, password)
        result["eval_user"] = eval_session.user
        assert_api_assertions(case.get("pre_api_assertions", []), eval_client, eval_session.token)

        start_url = build_frontend_url(args.eval_frontend_url, case.get("start_path", "/"))
        login_url = build_frontend_url(args.eval_frontend_url, "/login")
        run = run_rpa_case(
            case_id=case["id"],
            case=case,
            rpa_client=rpa_client,
            start_url=start_url,
            auth_token=eval_session.token,
            username=username,
            password=password,
            timeout_s=result["timeout_s"],
        )
        result["session_id"] = run.session_id
        result["raw_events"] = run.raw_events
        result["metrics"] = collect_metrics(run.raw_events, run.session)
        assert_metrics(case.get("assertions", {}), result["metrics"])
        assert_api_assertions(case.get("api_assertions", []), eval_client, eval_session.token)
        assert_expected_telemetry(case.get("expected", {}), result["metrics"])
        result["passed"] = True
    except RpaClawTimeoutError as exc:
        result["session_id"] = exc.session_id
        result["raw_events"] = exc.raw_events
        result["metrics"] = collect_metrics(exc.raw_events, None)
        result["failure_stage"] = classify_failure(exc)
        result["failure_message"] = str(exc)
    except (EvalAppError, RpaClawError, CaseAssertionError, AssertionError, KeyError) as exc:
        result["failure_stage"] = classify_failure(exc)
        result["failure_message"] = str(exc)
    except Exception as exc:
        result["failure_stage"] = "runner"
        result["failure_message"] = f"{type(exc).__name__}: {exc}"
    finally:
        result["latency_ms"] = round((time.perf_counter() - started) * 1000)
    return result


def run_rpa_case(
    *,
    case_id: str,
    case: dict[str, Any],
    rpa_client: RpaClawClient,
    start_url: str,
    auth_token: str,
    username: str,
    password: str,
    timeout_s: float,
) -> RpaRunResult:
    session_id = rpa_client.start_session(case_id)
    try:
        rpa_client.navigate(session_id, build_eval_auth_url(start_url, auth_token))
        rpa_client.navigate(session_id, start_url)
        business_instruction = build_browser_instruction(
            case=case,
            login_url="",
            start_url=start_url,
            username=username,
            password=password,
        )
        business_events = rpa_client.chat_with_wall_timeout(session_id, business_instruction, timeout_s=timeout_s)

        session = rpa_client.get_session(session_id)
        return RpaRunResult(session_id=session_id, raw_events=tag_events(business_events, "case"), session=session)
    finally:
        rpa_client.stop_session(session_id, ignore_errors=True)


def tag_events(events: list[dict[str, Any]], phase: str) -> list[dict[str, Any]]:
    return [{**event, "phase": phase} for event in events]


def resolve_case_timeout_s(case: dict[str, Any], args: argparse.Namespace) -> float:
    raw_value = case.get("timeout_s", args.case_timeout_s)
    try:
        timeout_s = float(raw_value)
    except (TypeError, ValueError) as exc:
        raise CaseAssertionError("case_data", f"invalid timeout_s for {case.get('id', '<unknown>')}: {raw_value!r}") from exc
    if timeout_s <= 0:
        raise CaseAssertionError("case_data", f"timeout_s must be positive for {case.get('id', '<unknown>')}")
    return timeout_s


def build_frontend_url(base_url: str, path: str) -> str:
    normalized_path = path if path.startswith("/") else f"/{path}"
    return f"{base_url.rstrip('/')}{normalized_path}"


def build_eval_auth_url(frontend_url: str, token: str) -> str:
    parsed = urlparse(frontend_url)
    base_url = f"{parsed.scheme}://{parsed.netloc}" if parsed.scheme and parsed.netloc else frontend_url
    return f"{base_url.rstrip('/')}/eval-auth.html?token={quote(token, safe='')}"


def build_browser_instruction(
    *,
    case: dict[str, Any],
    login_url: str,
    start_url: str,
    username: str,
    password: str,
) -> str:
    return "\n".join(
        [
            "你正在执行 RPA 评测用例。系统已经完成登录，并已导航到用例起始页面。",
            f"当前起始页面应为：{start_url}",
            "请只执行下面的业务任务，不要重新登录，不要把打开当前页面或返回当前 URL 当作完成。",
            "只有业务目标已经通过浏览器真实完成后，才能结束本次 recording command。",
            case["instruction"],
        ]
    )


def collect_metrics(raw_events: list[dict[str, Any]], session: dict[str, Any] | None) -> dict[str, Any]:
    trace_events = [event for event in raw_events if event.get("event") in {"trace_added", "agent_step_done"}]
    diagnostics = []
    for event in raw_events:
        data = event.get("data")
        if isinstance(data, dict):
            if isinstance(data.get("diagnostics"), list):
                diagnostics.extend(data["diagnostics"])
        if event.get("event") == "error":
            diagnostics.append(data if data is not None else event)

    session_traces = []
    if isinstance(session, dict):
        session_traces = session.get("traces") or []
        runtime_results = session.get("runtime_results") or {}
    else:
        runtime_results = {}

    final_url = extract_final_url(raw_events, session)
    visible_text = extract_visible_text(raw_events, session)
    downloads = extract_downloads(raw_events, session)

    return {
        "accepted_trace_count": max(len(trace_events), len(session_traces)),
        "diagnostics_count": len(diagnostics),
        "event_count": len(raw_events),
        "runtime_result_count": len(runtime_results.get("values", [])) if isinstance(runtime_results, dict) else 0,
        "final_url": final_url,
        "visible_text": visible_text,
        "output_text": extract_output_text(raw_events, session),
        "downloads": downloads,
    }


def assert_metrics(assertions: dict[str, Any], metrics: dict[str, Any]) -> None:
    accepted_min = assertions.get("accepted_trace_min")
    if accepted_min is not None and metrics["accepted_trace_count"] < accepted_min:
        raise CaseAssertionError(
            "assertion",
            f"accepted_trace_count {metrics['accepted_trace_count']} is below required minimum {accepted_min}"
        )
    diagnostics_max = assertions.get("diagnostics_max")
    if diagnostics_max is not None and metrics["diagnostics_count"] > diagnostics_max:
        raise CaseAssertionError(
            "assertion",
            f"diagnostics_count {metrics['diagnostics_count']} exceeds allowed maximum {diagnostics_max}"
        )


def assert_expected_telemetry(expected: dict[str, Any], metrics: dict[str, Any]) -> None:
    expected_path = expected.get("final_url_path")
    if expected_path:
        final_url = metrics.get("final_url")
        if not final_url:
            raise CaseAssertionError("unsupported_telemetry", "final_url_path assertion requires final URL telemetry")
        actual_path = urlparse(final_url).path if "://" in final_url else final_url
        if actual_path.rstrip("/") != str(expected_path).rstrip("/"):
            raise CaseAssertionError("assertion", f"expected final URL path {expected_path}, got {actual_path}")

    extracted_fields = expected.get("extracted_fields") or {}
    visible_text = expected.get("visible_text") or []
    if visible_text:
        text_blob = metrics.get("visible_text") or ""
        if not text_blob:
            raise CaseAssertionError("unsupported_telemetry", "visible_text assertions require page text telemetry")
        missing = [value for value in visible_text if not text_contains_expected(text_blob, value)]
        if missing:
            raise CaseAssertionError("assertion", f"expected text values were not found: {missing}")

    expected_output_text = expected.get("output_text") or []
    if expected_output_text:
        output_blob = metrics.get("output_text") or ""
        if not output_blob:
            raise CaseAssertionError("unsupported_output_telemetry", "output_text assertions require agent output telemetry")
        missing = [value for value in expected_output_text if not text_contains_expected(output_blob, value)]
        if missing:
            raise CaseAssertionError("assertion", f"expected output text values were not found: {missing}")

    if extracted_fields:
        output_blob = metrics.get("output_text") or ""
        if not output_blob:
            raise CaseAssertionError("unsupported_output_telemetry", "extracted_fields assertions require agent output telemetry")
        missing = [value for value in extracted_fields.values() if not text_contains_expected(output_blob, value)]
        if missing:
            raise CaseAssertionError("assertion", f"expected extracted field values were not found in agent output: {missing}")

    expected_download = expected.get("download") or {}
    expected_filename = expected_download.get("filename")
    expected_contains = [str(value) for value in expected_download.get("contains", [])]
    if expected_filename or expected_contains:
        downloads = metrics.get("downloads") or []
        if not downloads:
            raise CaseAssertionError("unsupported_telemetry", "download assertions require download telemetry")
        download_blob = "\n".join(str(item) for item in downloads)
        if expected_filename and str(expected_filename) not in download_blob:
            raise CaseAssertionError("assertion", f"expected download filename {expected_filename} was not found")
        if expected_contains:
            content = read_download_content(downloads, expected_filename)
            if content is None:
                raise CaseAssertionError(
                    "unsupported_download_content_telemetry",
                    "download contains assertion requires local download path or readable content telemetry",
                )
            missing = [value for value in expected_contains if value not in content]
            if missing:
                raise CaseAssertionError("assertion", f"expected download content values were not found: {missing}")


def assert_api_assertions(assertions: list[dict[str, Any]], eval_client: EvalAppClient, token: str) -> None:
    for assertion in assertions:
        path = assertion["path"]
        data = eval_client.get_json(path, token)
        target = find_api_target(data, assertion.get("find"))
        if assertion.get("absent"):
            if target is not None:
                raise CaseAssertionError(
                    "api_assertion",
                    f"{assertion.get('name', path)} target should be absent but was found",
                )
            continue
        if target is None:
            raise CaseAssertionError("api_assertion", f"{assertion.get('name', path)} target was not found")
        expected = assertion.get("expect") or {}
        for field, expected_value in expected.items():
            actual_value = target.get(field) if isinstance(target, dict) else None
            if not values_match(actual_value, expected_value):
                raise CaseAssertionError(
                    "api_assertion",
                    f"{assertion.get('name', path)} expected {field}={expected_value!r}, got {actual_value!r}",
                )


def find_api_target(data: Any, criteria: dict[str, Any] | None) -> dict[str, Any] | None:
    if criteria is None:
        return data if isinstance(data, dict) else None
    candidates = data if isinstance(data, list) else [data]
    for candidate in candidates:
        if not isinstance(candidate, dict):
            continue
        if all(candidate.get(key) == value for key, value in criteria.items()):
            return candidate
    return None


def values_match(actual: Any, expected: Any) -> bool:
    if isinstance(expected, dict):
        if not isinstance(actual, dict):
            return False
        return all(values_match(actual.get(key), value) for key, value in expected.items())
    if isinstance(expected, list):
        if not isinstance(actual, list):
            return False
        return all(any(values_match(item, expected_item) for item in actual) for expected_item in expected)
    if is_number_like(actual) and is_number_like(expected):
        return decimal_number(actual) == decimal_number(expected)
    return actual == expected


def extract_final_url(raw_events: list[dict[str, Any]], session: dict[str, Any] | None) -> str | None:
    for trace in reversed(extract_accepted_traces(raw_events, session)):
        after_page = trace.get("after_page")
        if isinstance(after_page, dict) and after_page.get("url"):
            return str(after_page["url"])

    for trace in reversed(extract_accepted_traces(raw_events, session)):
        for source in (trace.get("output"), trace):
            urls = extract_keyed_strings(source, {"final_url", "current_url", "opened_url", "url"})
            if urls:
                return urls[-1]

    urls: list[str] = []
    for source in [*raw_events, session]:
        urls.extend(extract_keyed_strings(source, {"url", "current_url", "page_url", "final_url"}))
    return urls[-1] if urls else None


def extract_accepted_traces(raw_events: list[dict[str, Any]], session: dict[str, Any] | None) -> list[dict[str, Any]]:
    traces = []
    for event in raw_events:
        data = event.get("data")
        if not isinstance(data, dict):
            continue
        trace = data.get("trace") if isinstance(data.get("trace"), dict) else data
        if not isinstance(trace, dict):
            continue
        if "after_page" not in trace and "output" not in trace:
            continue
        if trace.get("accepted") is False:
            continue
        traces.append(trace)
    if isinstance(session, dict):
        for trace in session.get("traces") or []:
            if isinstance(trace, dict) and trace.get("accepted") is not False:
                traces.append(trace)
    return traces


def extract_visible_text(raw_events: list[dict[str, Any]], session: dict[str, Any] | None) -> str:
    text_keys = {"text", "visible_text", "page_text", "content", "markdown", "message"}
    chunks = []
    for source in [*raw_events, session]:
        chunks.extend(extract_keyed_strings(source, text_keys))
    return "\n".join(chunks)


def extract_output_text(raw_events: list[dict[str, Any]], session: dict[str, Any] | None) -> str:
    chunks: list[str] = []
    for trace in extract_accepted_traces(raw_events, session):
        if "output" in trace:
            chunks.extend(flatten_strings(trace["output"]))
    for event in raw_events:
        data = event.get("data")
        if isinstance(data, dict) and "output" in data:
            chunks.extend(flatten_strings(data["output"]))
    if isinstance(session, dict):
        runtime_results = session.get("runtime_results") or {}
        if isinstance(runtime_results, dict):
            chunks.extend(flatten_strings(runtime_results.get("values", [])))
    return "\n".join(chunks)


def extract_downloads(raw_events: list[dict[str, Any]], session: dict[str, Any] | None) -> list[str]:
    downloads = []
    for source in [*raw_events, session]:
        downloads.extend(extract_download_entries(source))
    return downloads


def extract_download_entries(value: Any) -> list[str]:
    found = []
    if isinstance(value, dict):
        for key, item in value.items():
            key_lower = str(key).lower()
            if key_lower in {"filename", "suggested_filename", "download_path"}:
                found.extend(flatten_strings(item))
            if "download" in key_lower:
                found.extend(extract_download_object_fields(item))
            found.extend(extract_download_entries(item))
    elif isinstance(value, list):
        for item in value:
            found.extend(extract_download_entries(item))
    return found


def extract_download_object_fields(value: Any) -> list[str]:
    return extract_keyed_strings(
        value,
        {"path", "file_path", "local_path", "filename", "suggested_filename", "content", "text", "body", "raw_content"},
    )


def read_download_content(downloads: list[str], expected_filename: str | None) -> str | None:
    inline_chunks = []
    candidate_paths = []
    for item in downloads:
        value = str(item)
        path = Path(value)
        if path.exists() and path.is_file():
            candidate_paths.append(path)
            continue
        if expected_filename and value.endswith(str(expected_filename)):
            path = Path(value)
            if path.exists() and path.is_file():
                candidate_paths.append(path)
                continue
        if value.strip() and value != expected_filename:
            inline_chunks.append(value)

    if candidate_paths:
        preferred = select_download_path(candidate_paths, expected_filename)
        content = read_artifact_text(preferred)
        if content is not None:
            return content
    return "\n".join(inline_chunks) if inline_chunks else None


def select_download_path(paths: list[Path], expected_filename: str | None) -> Path:
    if expected_filename:
        for path in paths:
            if path.name == expected_filename:
                return path
    return paths[0]


def read_artifact_text(path: Path) -> str | None:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".csv", ".md", ".log"}:
        return path.read_text(encoding="utf-8", errors="replace")
    if suffix == ".json":
        return json_values_text(path)
    if suffix == ".xlsx":
        return read_xlsx_text(path)
    return None


def json_values_text(path: Path) -> str:
    raw = path.read_text(encoding="utf-8", errors="replace")
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return raw
    return "\n".join(flatten_strings(parsed))


def read_xlsx_text(path: Path) -> str | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        values = []
        for sheet in workbook.worksheets:
            values.append(sheet.title)
            for row in sheet.iter_rows(values_only=True):
                values.extend(str(cell) for cell in row if cell is not None)
        return "\n".join(values)
    finally:
        workbook.close()


def extract_keyed_strings(value: Any, keys: set[str]) -> list[str]:
    found = []
    if isinstance(value, dict):
        for key, item in value.items():
            if key in keys:
                found.extend(flatten_strings(item))
            found.extend(extract_keyed_strings(item, keys))
    elif isinstance(value, list):
        for item in value:
            found.extend(extract_keyed_strings(item, keys))
    return found


def flatten_strings(value: Any) -> list[str]:
    if isinstance(value, str):
        return [value]
    if isinstance(value, (int, float, bool)):
        return [str(value)]
    if isinstance(value, dict):
        strings = []
        for item in value.values():
            strings.extend(flatten_strings(item))
        return strings
    if isinstance(value, list):
        strings = []
        for item in value:
            strings.extend(flatten_strings(item))
        return strings
    return []


def text_contains_expected(text: str, expected: Any) -> bool:
    expected_text = str(expected)
    if expected_text in text:
        return True
    if is_number_like(expected):
        expected_number = decimal_number(expected)
        return any(number == expected_number for number in extract_decimal_numbers(text))
    normalized_text = normalize_text_for_match(text)
    normalized_expected = normalize_text_for_match(expected_text)
    return bool(normalized_expected and normalized_expected in normalized_text)


def is_number_like(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float, Decimal)):
        return True
    if isinstance(value, str):
        return bool(re.fullmatch(r"[+-]?\d+(?:\.\d+)?", value.strip()))
    return False


def decimal_number(value: Any) -> Decimal:
    return Decimal(str(value).replace(",", "")).normalize()


def extract_decimal_numbers(text: str) -> list[Decimal]:
    numbers = []
    for raw in re.findall(r"[+-]?\d[\d,]*(?:\.\d+)?", text):
        try:
            numbers.append(Decimal(raw.replace(",", "")).normalize())
        except InvalidOperation:
            continue
    return numbers


def normalize_text_for_match(value: str) -> str:
    return re.sub(r"\s+", "", value).replace(",", "").replace("¥", "").replace("￥", "")


def render_console_summary(case_results: list[dict[str, Any]]) -> str:
    total = len(case_results)
    passed = sum(1 for case in case_results if case.get("passed"))
    failed = total - passed
    lines = [
        "Evaluation Summary",
        f"Total: {total}  Passed: {passed}  Failed: {failed}",
        "",
        "| Case | Result | Latency | Failure |",
        "| --- | --- | ---: | --- |",
    ]
    for case in case_results:
        result = "PASS" if case.get("passed") else "FAIL"
        latency = case.get("latency_ms") or 0
        failure = ""
        if not case.get("passed"):
            failure = f"{case.get('failure_stage') or ''}: {case.get('failure_message') or ''}".strip(": ")
        lines.append(f"| {case['id']} | {result} | {latency} ms | {failure} |")
    return "\n".join(lines)


def classify_failure(exc: Exception) -> str:
    if isinstance(exc, RpaClawTimeoutError):
        return "timeout"
    if isinstance(exc, CaseAssertionError):
        return exc.stage
    if isinstance(exc, EvalAppError):
        return "eval_app"
    if isinstance(exc, RpaClawError):
        return "rpaclaw"
    if isinstance(exc, AssertionError):
        return "assertion"
    if isinstance(exc, KeyError):
        return "case_data"
    return "runner"


if __name__ == "__main__":
    raise SystemExit(main())
